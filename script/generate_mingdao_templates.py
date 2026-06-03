"""Generate 6 Excel templates for Mingdao Cloud import with typed sample rows."""

from __future__ import annotations

import datetime as dt
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "templates" / "mingdao-import"
OUT.mkdir(parents=True, exist_ok=True)

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TYPE_FILL = PatternFill("solid", fgColor="D9E1F2")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
SAMPLE_NOTE = "（示例行，帮助识别数值/日期；导入建表后可删）"

# excel_kind: int | decimal1 | decimal2 | date | text | url | single | member | longtext
TABLES = [
    {
        "file": "01-SEO自动数据看板.xlsx",
        "upsert": "日期",
        "fields": [
            ("日期", "日期", "脚本", "", "唯一键", "date", dt.date(2026, 5, 27)),
            ("自然点击", "数值", "GSC", "", "0位小数", "int", 128),
            ("展示量", "数值", "GSC", "", "0位小数", "int", 5420),
            ("平均CTR", "数值", "GSC", "", "2位小数；或改百分比", "decimal2", 0.052),
            ("全站加权平均排名", "数值", "GSC", "", "1位小数", "decimal1", 18.6),
            ("Top10词数", "数值", "Ahrefs+脚本", "", "0位", "int", 5),
            ("Top30词数", "数值", "Ahrefs+脚本", "", "0位", "int", 12),
            ("Backlinks变化", "数值", "Ahrefs", "", "近7天RD净值；0位可负", "int", 2),
            ("周平均排名", "数值", "脚本", "", "7日排名均值；1位小数", "decimal1", 20.5),
            ("周自然点击", "数值", "脚本", "", "7日点击均值；1位小数", "decimal1", 42.3),
            ("已监控URL收录数", "数值", "GSC+脚本", "", "0位", "int", 8),
            ("已监控URL异常数", "数值", "GSC+脚本", "", "0位", "int", 1),
            ("异常预警", "单选", "脚本", "正常|流量下跌|收录异常", "导入后改单选", "single", "正常"),
            ("周环比流量", "数值", "脚本", "", "1位；或百分比", "decimal1", 0.15),
            ("周环比Top30词", "数值", "脚本", "", "1位；或百分比", "decimal1", 0.08),
        ],
    },
    {
        "file": "02-站点关键词库.xlsx",
        "upsert": "关键词",
        "fields": [
            ("关键词", "文本", "API导入", "", "唯一键", "text", "cnc machining"),
            ("搜索量", "数值", "API导入", "", "0位", "int", 1200),
            ("KD", "数值", "API导入", "", "0位", "int", 45),
            ("当前排名", "数值", "API导入", "", "0位", "int", 23),
            ("排名变化", "文本", "API导入", "", "", "text", "↑3"),
            ("排名落地页", "链接", "API导入", "", "导入后改链接", "url", "https://www.cncpioneer.com/"),
            ("数据日期", "日期", "脚本", "", "", "date", dt.date(2026, 5, 27)),
            ("优先级", "单选", "人工", "P0|P1|P2|未分级", "导入后改单选", "single", "P0"),
            ("关键词类型", "单选", "人工", "产品词|工艺词|应用词|材质词|询盘词", "导入后改单选", "single", "询盘词"),
            ("优化状态", "单选", "人工", "未开始|优化中|已上Top30|已上Top10|停滞", "导入后改单选", "single", "优化中"),
            ("关联落地页", "链接", "人工", "", "导入后改链接", "url", "https://www.cncpioneer.com/"),
        ],
    },
    {
        "file": "03-页面管理表.xlsx",
        "upsert": "页面URL",
        "fields": [
            ("页面URL", "链接", "人工", "", "唯一键", "url", "https://www.cncpioneer.com/"),
            ("页面类型", "单选", "人工", "产品页|工艺页|案例页|服务页|询盘页|首页", "导入后改单选", "single", "首页"),
            ("关联关键词", "文本", "人工", "", "", "text", "cnc machining"),
            ("发布或优化日期", "日期", "人工", "", "", "date", dt.date(2026, 5, 1)),
            ("优化记录", "多行文本", "人工", "", "导入后改多行文本", "longtext", "更新H1与首段"),
            ("收录状态", "单选", "GSC", "已收录|未收录|索引异常", "导入后改单选", "single", "已收录"),
            ("页面流量", "数值", "GSC", "", "0位", "int", 15),
            ("数据日期", "日期", "脚本", "", "", "date", dt.date(2026, 5, 27)),
            ("质量校验", "单选", "人工", "合格|待整改|不合格", "导入后改单选", "single", "合格"),
        ],
    },
    {
        "file": "04-外链监控表.xlsx",
        "upsert": "来源URL + 目标URL",
        "fields": [
            ("来源域名", "文本", "API导入", "", "", "text", "example.com"),
            ("来源URL", "链接", "API导入", "", "组合唯一键", "url", "https://example.com/article/1"),
            ("目标URL", "链接", "API导入", "", "组合唯一键", "url", "https://www.cncpioneer.com/"),
            ("锚文本", "文本", "API导入", "", "", "text", "CNC machining"),
            ("域名DR", "数值", "API导入", "", "0位", "int", 45),
            ("是否Dofollow", "单选", "API导入", "是|否", "导入后改单选", "single", "是"),
            ("首次发现", "日期", "API导入", "", "", "date", dt.date(2025, 11, 3)),
            ("外链存活", "单选", "API导入", "有效|已失效", "导入后改单选", "single", "有效"),
            ("数据日期", "日期", "脚本", "", "", "date", dt.date(2026, 5, 27)),
            ("行业相关性", "单选", "人工", "高度相关|一般|不相关", "导入后改单选", "single", "高度相关"),
            ("审核结果", "单选", "人工", "合格|不合格|待审核", "导入后改单选", "single", "待审核"),
            ("违规说明", "多行文本", "人工", "", "导入后改多行文本", "longtext", ""),
        ],
    },
    {
        "file": "05-SEO每日执行日志.xlsx",
        "upsert": "日期 + 运营人",
        "fields": [
            ("日期", "日期", "人工", "", "组合唯一键", "date", dt.date(2026, 5, 27)),
            ("运营人", "成员", "人工", "", "导入后改成员", "member", "张三"),
            ("今日主攻词", "文本", "人工", "", "", "text", "cnc machining"),
            ("新增页面URL", "链接", "人工", "", "导入后改链接", "url", ""),
            ("优化页面URL", "链接", "人工", "", "导入后改链接", "url", "https://www.cncpioneer.com/"),
            ("今日有效动作", "多行文本", "人工", "", "导入后改多行文本", "longtext", "优化首页标题"),
            ("今日无效动作", "多行文本", "人工", "", "导入后改多行文本", "longtext", "堆无关关键词"),
            ("数据波动判断", "多行文本", "人工", "", "导入后改多行文本", "longtext", ""),
            ("明日攻坚任务", "多行文本", "人工", "", "导入后改多行文本", "longtext", "推进P0词落地页"),
            ("当日合规", "单选", "人工", "达标|未达标", "导入后改单选", "single", "达标"),
        ],
    },
    {
        "file": "06-周绩效复盘.xlsx",
        "upsert": "周次 + 运营人",
        "fields": [
            ("周次", "文本", "人工", "", "如2026-W22", "text", "2026-W22"),
            ("运营人", "成员", "人工", "", "导入后改成员", "member", "张三"),
            ("本周Top30净增", "数值", "人工", "", "0位", "int", 2),
            ("本周合格RD数", "数值", "人工", "", "0位", "int", 3),
            ("本周新增页面数", "数值", "人工", "", "0位", "int", 1),
            ("本周优化页面数", "数值", "人工", "", "0位", "int", 2),
            ("本周有效动作总结", "多行文本", "人工", "", "导入后改多行文本", "longtext", "动作1"),
            ("本周无效动作总结", "多行文本", "人工", "", "导入后改多行文本", "longtext", "动作1"),
            ("问题根因分析", "多行文本", "人工", "", "导入后改多行文本", "longtext", ""),
            ("下周KPI", "多行文本", "人工", "", "导入后改多行文本", "longtext", "Top30+3"),
            ("周考核等级", "单选", "人工", "优秀|合格|不合格", "导入后改单选", "single", "合格"),
        ],
    },
]

FORMATS = {
    "int": "0",
    "decimal1": "0.0",
    "decimal2": "0.00",
    "date": "yyyy-mm-dd",
    "text": "@",
    "url": "@",
    "single": "@",
    "member": "@",
    "longtext": "@",
}


def style_header(ws, row: int, cols: int) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def apply_typed_cell(cell, kind: str, value) -> None:
    fmt = FORMATS[kind]
    cell.number_format = fmt
    if value == "" or value is None:
        cell.value = None
        return
    if kind == "date":
        cell.value = value
        return
    if kind in {"int", "decimal1", "decimal2"}:
        cell.value = value
        return
    cell.value = str(value)


def auto_width(ws, max_col: int) -> None:
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        max_len = 12
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value is not None and cell.value != "":
                    max_len = max(max_len, min(len(str(cell.value)) + 2, 42))
        ws.column_dimensions[letter].width = max_len


def build_type_row(fields) -> list[str]:
    return [f[1] for f in fields]


def write_csv_guide(all_rows: list[list[str]]) -> None:
    path = OUT / "字段类型对照-导入后必改.csv"
    header = ["工作表", "字段名", "明道云控件", "Excel示例格式", "单选选项", "导入后是否还要改"]
    lines = [",".join(header)]
    for row in all_rows:
        lines.append(",".join(f'"{cell}"' for cell in row))
    path.write_text("\n".join(lines), encoding="utf-8-sig")


def main() -> None:
    csv_rows: list[list[str]] = []

    for table in TABLES:
        wb = Workbook()
        ws_data = wb.active
        ws_data.title = "数据"

        fields = table["fields"]
        headers = [f[0] for f in fields]
        types = build_type_row(fields)

        ws_data.append(headers)
        style_header(ws_data, 1, len(headers))

        ws_data.append(types)
        for col in range(1, len(types) + 1):
            cell = ws_data.cell(row=2, column=col)
            cell.fill = TYPE_FILL
            cell.font = Font(italic=True, color="1F4E79")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        sample_values = [f[6] for f in fields]
        ws_data.append(sample_values)
        for col, field in enumerate(fields, start=1):
            apply_typed_cell(ws_data.cell(row=3, column=col), field[5], field[6])

        ws_data.freeze_panes = "A4"
        auto_width(ws_data, len(headers))

        ws_meta = wb.create_sheet("字段说明")
        meta_headers = ["字段名", "明道云控件", "数据来源", "单选选项", "小数位建议", "说明"]
        ws_meta.append(meta_headers)
        style_header(ws_meta, 1, len(meta_headers))
        ws_meta.append(["【Upsert唯一键】", table["upsert"], "", "", "", "导入后设必填/不可重复"])
        for cell in ws_meta[2]:
            cell.fill = NOTE_FILL

        for name, mtype, source, options, note, kind, _sample in fields:
            decimals = {
                "int": "0位",
                "decimal1": "1位",
                "decimal2": "2位",
            }.get(kind, "")
            must_fix = "否" if mtype == "数值" and kind in {"int", "decimal1", "decimal2"} else (
                "否" if mtype == "日期" else "是（改控件+选项）"
            )
            ws_meta.append([name, mtype, source, options, decimals, note or must_fix])
            csv_rows.append([
                table["file"],
                name,
                mtype,
                kind,
                options,
                must_fix,
            ])
        auto_width(ws_meta, 6)

        ws_guide = wb.create_sheet("导入说明")
        guide_lines = [
            ["明道云 Excel 导入说明"],
            [""],
            ["【重要】明道云不能 100% 从 Excel 读出控件类型。本文件已做两件事："],
            ["  1) 第2行 = 明道云控件类型参考（导入时若出现「类型映射」步骤，请按此行选择）"],
            ["  2) 第3行 = 带 Excel 单元格格式的示例（数值/日期可被识别；导入建表后可删除第2-3行）"],
            [""],
            ["导入步骤："],
            ["  1. 明道云 → 新建工作表 → 从 Excel 导入 → 选「数据」表"],
            ["  2. 若有「字段类型映射」，按第2行把列改成：数值/日期/单选/链接/多行文本/成员"],
            ["  3. 单选字段：导入后进入字段设置，添加选项（必须与「字段说明」表完全一致）"],
            ["  4. 链接、成员、多行文本：Excel 无法自动识别，导入后手动改控件类型"],
            ["  5. 数值字段小数位：整数0位；排名1位；CTR/环比2位或1位（见 docs/dashboard.md）"],
            ["  6. 重新生成：python script/generate_mingdao_templates.py"],
            [""],
            ["建表完成后：可只保留第1行表头，删除第2-3行示例数据。"],
        ]
        for line in guide_lines:
            ws_guide.append(line)
        ws_guide["A1"].font = Font(bold=True, size=12)
        auto_width(ws_guide, 1)

        path = OUT / table["file"]
        wb.save(path)
        print(f"Created {path}")

    write_csv_guide(csv_rows)
    print(f"Created {OUT / '字段类型对照-导入后必改.csv'}")


if __name__ == "__main__":
    main()
